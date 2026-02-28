"""
Flask后端服务 - PubMed文献搜索与AI总结
提供REST API接口供前端调用
"""

import os
import uuid
import threading
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import sys

# 添加父目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import config
from pubmed_crawler import PubMedCrawler
from journal_filter import JournalFilter
from summarizer import ArticleSummarizer

app = Flask(__name__, static_folder='static', static_url_path='/static')
CORS(app)

# 任务存储
tasks = {}

# 期刊配置
journals_config = config.JOURNALS.copy()

# 输出目录使用绝对路径
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), config.OUTPUT_DIR)

# 确保输出目录存在
os.makedirs(OUTPUT_DIR, exist_ok=True)


# 前端页面路由
@app.route('/')
def index():
    """渲染前端页面"""
    return send_from_directory('static', 'index.html')


def save_markdown_report(articles, output_path, overall_summary=""):
    """保存Markdown报告"""
    with open(output_path, "w", encoding="utf-8") as f:
        if overall_summary:
            f.write(overall_summary)
            f.write("\n\n")
        f.write("---\n\n")
        for i, article in enumerate(articles, 1):
            f.write(f"## 文章 {i}: {article.get('title', '无标题')}\n\n")
            f.write(f"**期刊**: {article.get('journal', 'N/A')}\n\n")
            f.write(f"**出版社**: {article.get('publisher', 'N/A')}\n\n")
            f.write(f"**发表日期**: {article.get('pub_date', 'N/A')}\n\n")
            f.write(f"**DOI**: {article.get('doi', 'N/A')}\n\n")
            authors = article.get('authors', '')
            if authors:
                f.write(f"**作者**: {authors}\n\n")
            abstract = article.get('abstract', '')
            if abstract:
                f.write("**摘要**:\n\n")
                f.write(f"{abstract}\n\n")
            summary = article.get('summary', '')
            if summary and summary != "无摘要":
                f.write("**AI总结**:\n\n")
                f.write(f"{summary}\n\n")
            f.write("---\n\n")


def save_excel(articles, output_path):
    """保存Excel文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "文献"

    headers = ["序号", "标题", "期刊", "出版社", "发表日期", "DOI", "作者", "摘要", "AI总结"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row, article in enumerate(articles, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=article.get("title", ""))
        ws.cell(row=row, column=3, value=article.get("journal", ""))
        ws.cell(row=row, column=4, value=article.get("publisher", ""))
        ws.cell(row=row, column=5, value=article.get("pub_date", ""))
        ws.cell(row=row, column=6, value=article.get("doi", ""))
        ws.cell(row=row, column=7, value=article.get("authors", ""))
        ws.cell(row=row, column=8, value=article.get("abstract", ""))
        ws.cell(row=row, column=9, value=article.get("summary", ""))

    column_widths = {1: 6, 2: 60, 3: 30, 4: 10, 5: 12, 6: 25, 7: 40, 8: 80, 9: 100}
    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    wb.save(output_path)


def run_search_task(task_id, params):
    """后台执行搜索任务"""
    task = tasks[task_id]
    task['status'] = 'running'
    task['progress'] = 0
    task['message'] = '正在初始化...'

    # 检查暂停/取消状态的辅助函数
    def check_pause():
        while task.get('paused', False) and not task.get('cancelled', False):
            import time
            time.sleep(0.5)
        if task.get('cancelled', False):
            raise Exception("任务已取消")

    try:
        from datetime import datetime
        os.makedirs(OUTPUT_DIR, exist_ok=True)

        # 步骤1: AI优化检索词
        check_pause()
        task['progress'] = 5
        task['message'] = 'AI优化检索词...'
        # 获取用户提供的API密钥，如果没有则使用默认配置
        api_key = params.get('api_key', config.DEEPSEEK_API_KEY)
        summarizer = ArticleSummarizer(api_key=api_key)
        user_topic = params['topic']
        optimized_terms = summarizer.optimize_search_terms(user_topic)

        # 步骤2: 搜索文章
        check_pause()
        task['progress'] = 10
        task['message'] = '正在搜索PubMed...'
        crawler = PubMedCrawler()
        start_date = params.get('start_date', '2025/01/01')
        end_date = params.get('end_date', datetime.now().strftime("%Y/%m/%d"))
        max_results = params.get('max_results', 30)

        all_articles = crawler.get_articles(
            search_terms=optimized_terms,
            start_date=start_date,
            end_date=end_date,
            max_results=max_results
        )

        if not all_articles:
            task['status'] = 'completed'
            task['progress'] = 100
            task['message'] = '未找到相关文章'
            task['results'] = []
            return

        # 步骤3: 筛选期刊
        check_pause()
        enable_filter = params.get('enable_filter', True)

        if enable_filter:
            task['progress'] = 30
            task['message'] = '筛选期刊...'
            journal_filter = JournalFilter()

            # 获取前端传递的期刊列表，如果没有则使用全部期刊
            selected_journals = params.get('selected_journals', [])
            if selected_journals:
                # 使用自定义期刊列表筛选
                filtered_articles = journal_filter.filter_articles_by_journals(all_articles, selected_journals)
            else:
                # 使用默认配置筛选
                filtered_articles = journal_filter.filter_articles(all_articles)
        else:
            # 不筛选，返回所有文章
            task['progress'] = 30
            task['message'] = '跳过筛选...'
            filtered_articles = all_articles
            for article in filtered_articles:
                article["publisher"] = "Unknown"

        if not filtered_articles:
            task['status'] = 'completed'
            task['progress'] = 100
            task['message'] = '筛选后没有符合条件的文章'
            task['results'] = []
            return

        # 步骤4: AI总结
        task['progress'] = 50
        task['message'] = 'AI总结文章中...'
        max_workers = params.get('max_workers', 5)

        # 使用字典来存储已完成文章的PMID和内容的映射
        completed_articles = {}

        # 创建进度回调函数，实时更新任务状态
        def progress_callback(article, completed, total):
            # 检查暂停或取消状态
            while task.get('paused', False) and not task.get('cancelled', False):
                import time
                time.sleep(0.5)

            if task.get('cancelled', False):
                return

            # 使用PMID或标题作为key
            key = article.get('pmid') or article.get('title', '')
            completed_articles[key] = article
            # 从原始列表中找出已完成的文章
            current_results = []
            for a in filtered_articles:
                k = a.get('pmid') or a.get('title', '')
                if k in completed_articles:
                    current_results.append(completed_articles[k])
                else:
                    current_results.append(a)
            task['results'] = current_results
            task['progress'] = 50 + int(30 * completed / total)
            task['message'] = f'AI总结文章中... ({completed}/{total})'

        summarized_articles = summarizer.summarize_articles(
            filtered_articles,
            max_workers=max_workers,
            progress_callback=progress_callback
        )

        # 步骤5: AI润色主题
        check_pause()
        task['progress'] = 80
        task['message'] = '生成文献综述...'
        polished_topic = summarizer.polish_search_topic(user_topic)

        # 步骤6: 生成文献综述
        check_pause()
        literature_review = summarizer.generate_literature_review(
            summarized_articles,
            polished_topic,
            start_date,
            end_date
        )

        # 步骤7: 保存文件
        check_pause()
        task['progress'] = 95
        task['message'] = '保存文件...'

        report_path = os.path.join(OUTPUT_DIR, f"{task_id}_report.md")
        review_path = os.path.join(OUTPUT_DIR, f"{task_id}_review.md")
        excel_path = os.path.join(OUTPUT_DIR, f"{task_id}_articles.xlsx")

        overall_summary = summarizer.generate_overall_summary(summarized_articles)
        save_markdown_report(summarized_articles, report_path, overall_summary)
        save_excel(summarized_articles, excel_path)

        with open(review_path, "w", encoding="utf-8") as f:
            f.write(literature_review)

        # 读取文献综述内容
        with open(review_path, "r", encoding="utf-8") as f:
            review_content = f.read()

        task['status'] = 'completed'
        task['progress'] = 100
        task['message'] = '完成!'
        task['results'] = summarized_articles
        task['files'] = {
            'report': os.path.basename(report_path),
            'review': os.path.basename(review_path),
            'excel': os.path.basename(excel_path)
        }
        task['review_content'] = review_content
        task['polished_topic'] = polished_topic

    except Exception as e:
        if task.get('cancelled', False):
            task['status'] = 'cancelled'
            task['message'] = '任务已取消'
        else:
            task['status'] = 'error'
            task['message'] = f'错误: {str(e)}'
            task['error'] = str(e)


# ========== API接口 ==========

@app.route('/api/config', methods=['GET'])
def get_config():
    """获取当前配置"""
    return jsonify({
        'journals': journals_config,
        'defaults': {
            'max_results': config.MAX_SEARCH_RESULTS,
            'max_workers': config.MAX_WORKERS,
            'start_date': '2025/01/01'
        }
    })


@app.route('/api/journals', methods=['POST'])
def update_journals():
    """更新期刊配置"""
    global journals_config
    journals_config = request.json.get('journals', journals_config)
    config.JOURNALS = journals_config
    return jsonify({'status': 'success'})


@app.route('/api/search', methods=['POST'])
def start_search():
    """启动搜索任务"""
    params = request.json

    task_id = str(uuid.uuid4())
    tasks[task_id] = {
        'status': 'pending',
        'progress': 0,
        'message': '等待中...',
        'params': params,
        'results': [],
        'files': {},
        'paused': False,
        'cancelled': False
    }

    # 后台启动任务
    thread = threading.Thread(target=run_search_task, args=(task_id, params))
    thread.start()

    return jsonify({
        'task_id': task_id,
        'status': 'pending'
    })


@app.route('/api/task/<task_id>', methods=['GET'])
def get_task_status(task_id):
    """获取任务状态"""
    task = tasks.get(task_id)
    if not task:
        return jsonify({'error': '任务不存在'}), 404

    response = {
        'status': task['status'],
        'progress': task['progress'],
        'message': task['message'],
        'result_count': len(task.get('results', [])),
        'paused': task.get('paused', False)
    }

    # 如果任务正在运行或已完成，返回当前结果供实时显示
    if task['status'] in ['running', 'completed', 'paused']:
        response['results'] = task.get('results', [])
        response['review_content'] = task.get('review_content', '')

    return jsonify(response)


@app.route('/api/task/<task_id>/pause', methods=['POST'])
def pause_task(task_id):
    """暂停任务"""
    task = tasks.get(task_id)
    if not task:
        return jsonify({'error': '任务不存在'}), 404

    if task['status'] != 'running':
        return jsonify({'error': '任务不在运行中'}), 400

    task['paused'] = True
    task['status'] = 'paused'
    task['message'] = '已暂停'

    return jsonify({'status': 'success', 'message': '任务已暂停'})


@app.route('/api/task/<task_id>/resume', methods=['POST'])
def resume_task(task_id):
    """恢复任务"""
    task = tasks.get(task_id)
    if not task:
        return jsonify({'error': '任务不存在'}), 404

    if task['status'] != 'paused':
        return jsonify({'error': '任务不在暂停状态'}), 400

    task['paused'] = False
    task['status'] = 'running'
    task['message'] = '继续运行...'

    return jsonify({'status': 'success', 'message': '任务已恢复'})


@app.route('/api/task/<task_id>/cancel', methods=['POST'])
def cancel_task(task_id):
    """取消任务"""
    task = tasks.get(task_id)
    if not task:
        return jsonify({'error': '任务不存在'}), 404

    task['cancelled'] = True
    task['status'] = 'cancelled'
    task['message'] = '任务已取消'

    return jsonify({'status': 'success', 'message': '任务已取消'})


@app.route('/api/task/<task_id>/results', methods=['GET'])
def get_task_results(task_id):
    """获取任务结果"""
    task = tasks.get(task_id)
    if not task:
        return jsonify({'error': '任务不存在'}), 404

    if task['status'] != 'completed':
        return jsonify({'error': '任务未完成'}), 400

    return jsonify({
        'results': task.get('results', []),
        'files': task.get('files', {}),
        'review_content': task.get('review_content', ''),
        'polished_topic': task.get('polished_topic', '')
    })


@app.route('/api/files/<path:filename>')
def download_file(filename):
    """下载输出文件"""
    return send_from_directory(OUTPUT_DIR, filename)


if __name__ == '__main__':
    print("=" * 50)
    print("PubMed文献搜索Web服务")
    print("访问地址: http://localhost:5000")
    print("=" * 50)
    app.run(debug=True, host='0.0.0.0', port=5000)