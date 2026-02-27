"""
PubMed食管癌文献爬取与AI总结 - 主程序
"""

import os
import sys
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

import config
from pubmed_crawler import PubMedCrawler
from journal_filter import JournalFilter
from summarizer import ArticleSummarizer


def create_output_dir():
    """创建输出目录"""
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)


def save_markdown_report(articles: list, output_path: str, overall_summary: str = ""):
    """
    保存Markdown报告

    Args:
        articles: 文章列表
        output_path: 输出文件路径
        overall_summary: 整体总结
    """
    with open(output_path, "w", encoding="utf-8") as f:
        # 写入整体总结
        if overall_summary:
            f.write(overall_summary)
            f.write("\n\n")

        # 写入每篇文章的详细信息
        f.write("---\n\n")

        for i, article in enumerate(articles, 1):
            f.write(f"## 文章 {i}: {article.get('title', '无标题')}\n\n")

            f.write(f"**PMID**: {article.get('pmid', 'N/A')}\n\n")
            f.write(f"**期刊**: {article.get('journal', 'N/A')}\n\n")
            f.write(f"**出版社**: {article.get('publisher', 'N/A')}\n\n")
            f.write(f"**发表日期**: {article.get('pub_date', 'N/A')}\n\n")
            f.write(f"**DOI**: {article.get('doi', 'N/A')}\n\n")

            authors = article.get('authors', '')
            if authors:
                f.write(f"**作者**: {authors}\n\n")

            # 摘要
            abstract = article.get('abstract', '')
            if abstract:
                f.write("**摘要**:\n\n")
                f.write(f"{abstract}\n\n")

            # 总结
            summary = article.get('summary', '')
            if summary and summary != "无摘要":
                f.write("**AI总结**:\n\n")
                f.write(f"{summary}\n\n")

            f.write("---\n\n")

        # 写入时间戳
        f.write(f"\n*报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*\n")

    print(f"Markdown报告已保存到: {output_path}")


def save_excel(articles: list, output_path: str):
    """
    保存Excel文件

    Args:
        articles: 文章列表
        output_path: 输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "食管癌文献"

    # 定义表头
    headers = ["序号", "PMID", "标题", "期刊", "出版社", "发表日期", "DOI", "作者", "摘要", "AI总结"]

    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 写入数据
    for row, article in enumerate(articles, 2):
        ws.cell(row=row, column=1, value=row-1)  # 序号
        ws.cell(row=row, column=2, value=article.get("pmid", ""))
        ws.cell(row=row, column=3, value=article.get("title", ""))
        ws.cell(row=row, column=4, value=article.get("journal", ""))
        ws.cell(row=row, column=5, value=article.get("publisher", ""))
        ws.cell(row=row, column=6, value=article.get("pub_date", ""))
        ws.cell(row=row, column=7, value=article.get("doi", ""))
        ws.cell(row=row, column=8, value=article.get("authors", ""))
        ws.cell(row=row, column=9, value=article.get("abstract", ""))
        ws.cell(row=row, column=10, value=article.get("summary", ""))

    # 调整列宽
    column_widths = {
        1: 6,   # 序号
        2: 12,  # PMID
        3: 60,  # 标题
        4: 30,  # 期刊
        5: 10,  # 出版社
        6: 12,  # 发表日期
        7: 25,  # DOI
        8: 40,  # 作者
        9: 80,  # 摘要
        10: 100 # AI总结
    }

    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 保存
    wb.save(output_path)
    print(f"Excel文件已保存到: {output_path}")


def main():
    """主函数"""
    from datetime import datetime

    # 解析命令行参数
    parser = argparse.ArgumentParser(description="PubMed文献搜索与AI总结工具")
    parser.add_argument("-t", "--topic", type=str, default="", help="搜索主题")
    parser.add_argument("-s", "--start-date", type=str, default="", help="开始日期 (YYYY/MM/DD)")
    parser.add_argument("-e", "--end-date", type=str, default="", help="结束日期 (YYYY/MM/DD)")
    parser.add_argument("-m", "--max-results", type=int, default=0, help="最大搜索篇数")
    parser.add_argument("-w", "--workers", type=int, default=0, help="并发线程数")
    parser.add_argument("--interactive", action="store_true", help="交互式模式")

    args = parser.parse_args()

    print("=" * 60)
    print("PubMed文献搜索与AI总结工具")
    print("=" * 60)

    # 创建输出目录
    create_output_dir()

    # 判断模式：命令行参数模式 或 交互式模式
    use_cli_args = args.topic and not args.interactive

    if use_cli_args:
        # 命令行参数模式
        user_topic = args.topic
        start_date = args.start_date if args.start_date else "2025/09/01"
        end_date = args.end_date if args.end_date else datetime.now().strftime("%Y/%m/%d")
        max_results = args.max_results if args.max_results > 0 else config.MAX_SEARCH_RESULTS
        max_workers = args.workers if args.workers > 0 else config.MAX_WORKERS

        print(f"\n使用命令行参数:")
        print(f"  主题: {user_topic}")
        print(f"  时间范围: {start_date} 至 {end_date}")
        print(f"  最大篇数: {max_results}")
        print(f"  并发线程: {max_workers}")

        # 步骤1: AI优化搜索词
        print("\n[步骤1] AI优化检索词...")
        summarizer = ArticleSummarizer()
        optimized_terms = summarizer.optimize_search_terms(user_topic)

        print("\n优化后的检索词:")
        for i, term in enumerate(optimized_terms, 1):
            print(f"  {i}. {term}")
    else:
        # 交互式模式
        # 步骤0: 用户输入搜索主题
        print("\n[步骤0] 输入搜索主题")
        print("-" * 40)
        user_topic = input("请输入搜索主题 (例如: 食管癌免疫治疗): ").strip()

        if not user_topic:
            print("未输入搜索主题，使用默认主题: 食管癌")
            user_topic = "食管癌"

        # 步骤1: AI优化搜索词
        print("\n[步骤1] AI优化检索词...")
        summarizer = ArticleSummarizer()
        optimized_terms = summarizer.optimize_search_terms(user_topic)

        print("\n优化后的检索词:")
        for i, term in enumerate(optimized_terms, 1):
            print(f"  {i}. {term}")

        # 步骤2: 用户输入搜索参数
        print("\n[步骤2] 设置搜索参数")
        print("-" * 40)

        # 输入开始日期
        default_start_date = "2025/09/01"
        start_date_input = input(f"开始日期 (YYYY/MM/DD，默认: {default_start_date}): ").strip()
        start_date = start_date_input if start_date_input else default_start_date

        # 输入结束日期
        default_end_date = datetime.now().strftime("%Y/%m/%d")
        end_date_input = input(f"结束日期 (YYYY/MM/DD，默认: {default_end_date}): ").strip()
        end_date = end_date_input if end_date_input else default_end_date

        # 输入最大搜索篇数
        default_max = config.MAX_SEARCH_RESULTS
        max_input = input(f"最大搜索篇数 (默认: {default_max}): ").strip()
        max_results = int(max_input) if max_input.isdigit() else default_max

        # 输入并发线程数
        max_workers_input = input(f"并发线程数 (默认: {config.MAX_WORKERS}): ").strip()
        max_workers = int(max_workers_input) if max_workers_input.isdigit() else config.MAX_WORKERS

        print(f"\n搜索参数:")
        print(f"  主题: {user_topic}")
        print(f"  时间范围: {start_date} 至 {end_date}")
        print(f"  最大篇数: {max_results}")
        print(f"  并发线程: {max_workers}")

    # 步骤3: 从PubMed爬取文章
    print(f"\n[步骤3] 从PubMed搜索「{user_topic}」相关文章...")
    crawler = PubMedCrawler()
    all_articles = crawler.get_articles(
        search_terms=optimized_terms,
        start_date=start_date,
        end_date=end_date,
        max_results=max_results
    )

    if not all_articles:
        print("未找到相关文章，程序退出")
        return

    # 步骤4: 筛选目标期刊
    print("\n[步骤4] 按照出版社标准筛选期刊...")
    journal_filter = JournalFilter()
    filtered_articles = journal_filter.filter_articles(all_articles)

    if not filtered_articles:
        print("筛选后没有符合条件的文章，程序退出")
        return

    # 步骤5: 使用Deepseek API进行总结（多线程）
    print("\n[步骤5] 使用Deepseek API总结文章（多线程）...")

    # 根据模式获取并发数
    if args.topic and not args.interactive:
        # 命令行模式，使用预设值
        max_workers = args.workers if args.workers > 0 else config.MAX_WORKERS
        print(f"使用并发线程数: {max_workers}")
    else:
        # 交互式模式
        max_workers_input = input(f"并发线程数 (默认: {config.MAX_WORKERS}): ").strip()
        max_workers = int(max_workers_input) if max_workers_input.isdigit() else config.MAX_WORKERS

    # 生成整体统计
    overall_summary = summarizer.generate_overall_summary(filtered_articles)

    # 批量总结每篇文章（多线程）
    summarized_articles = summarizer.summarize_articles(filtered_articles, max_workers=max_workers)

    # 步骤6: AI润色搜索主题并生成文献综述
    print("\n[步骤6] AI润色搜索主题...")
    polished_topic = summarizer.polish_search_topic(user_topic)
    print(f"  原始主题: {user_topic}")
    print(f"  润色主题: {polished_topic}")

    print("\n[步骤7] 生成文献综述...")
    literature_review = summarizer.generate_literature_review(
        summarized_articles,
        polished_topic,
        start_date,
        end_date
    )

    # 步骤8: 保存输出文件
    print("\n[步骤8] 保存输出文件...")

    # 保存Markdown报告（不包含文献综述）
    report_path = os.path.join(config.OUTPUT_DIR, config.REPORT_FILE)
    save_markdown_report(summarized_articles, report_path, overall_summary)

    # 保存文献综述到单独文件
    review_path = os.path.join(config.OUTPUT_DIR, config.REVIEW_FILE)
    with open(review_path, "w", encoding="utf-8") as f:
        f.write(literature_review)
    print(f"文献综述已保存到: {review_path}")

    # 保存Excel文件
    excel_path = os.path.join(config.OUTPUT_DIR, config.EXCEL_FILE)
    save_excel(summarized_articles, excel_path)

    print("\n" + "=" * 60)
    print("完成!")
    print(f"找到 {len(summarized_articles)} 篇符合条件的文章")
    print(f"文章报告: {report_path}")
    print(f"文献综述: {review_path}")
    print(f"数据文件: {excel_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()